import json
from openpyxl import Workbook, load_workbook

class BudgetApp:
    def __init__(self):
        self.income = 0
        self.expenses = []
        self.budget = 0

    def get_income(self):
        self.income = float(input("Enter your monthly income: $"))

    def add_expense(self, category, amount):
        self.expenses.append({"category": category, "amount": amount})

    def calculate_budget(self):
        total_expenses = sum(expense["amount"] for expense in self.expenses)
        self.budget = self.income - total_expenses

    def display_budget(self):
        print("\nBudget Summary:")
        print(f"Income: ${self.format_decimal(self.income)}")
        print("Expenses:")
        for expense in self.expenses:
            print(f"  {expense['category']}: ${self.format_decimal(expense['amount'])}")
        print(f"Remaining Budget: ${self.format_decimal(self.budget)}\n")

    def format_decimal(self, value):
        # Format the value with exactly two decimal places
        return '{:.2f}'.format(value)

    def save_to_excel(self, filename='budget.xlsx'):
        workbook = Workbook()
        sheet = workbook.active

        # Write header
        sheet.append(['Category', 'Amount'])
        
        # Write expenses
        for expense in self.expenses:
            sheet.append([expense['category'], expense['amount']])
        
        # Write income and budget
        sheet.append(['Income', self.income])
        sheet.append(['Remaining Budget', self.budget])

        workbook.save(filename)
        print(f"Data saved to {filename}")

    def load_from_excel(self, filename='budget.xlsx'):
        try:
            workbook = load_workbook(filename)
            sheet = workbook.active

            self.expenses = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    self.expenses.append({"category": row[0], "amount": row[1]})

            self.income = sheet.cell(row=sheet.max_row - 1, column=2).value
            self.budget = sheet.cell(row=sheet.max_row, column=2).value

            print(f"Data loaded from {filename}")

        except FileNotFoundError:
            print(f"{filename} not found. Starting with a fresh budget.")

def main():
    budget_app = BudgetApp()

    # Load existing data from the file
    budget_app.load_from_excel()

    while True:
        print("Budget App Menu:")
        print("1. Set Income")
        print("2. Add Expense")
        print("3. Calculate Budget")
        print("4. Display Budget")
        print("5. Save to Excel")
        print("6. Exit")

        choice = input("Enter your choice (1-6): ")

        if choice == "1":
            budget_app.get_income()
        elif choice == "2":
            category = input("Enter expense category: ")
            amount = float(input("Enter expense amount: $"))
            budget_app.add_expense(category, amount)
        elif choice == "3":
            budget_app.calculate_budget()
            print("Budget calculated successfully.")
        elif choice == "4":
            budget_app.display_budget()
        elif choice == "5":
            budget_app.save_to_excel()
        elif choice == "6":
            print("Exiting Budget App. Goodbye!")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 6.")

if __name__ == "__main__":
    main()
